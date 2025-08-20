from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# from step4_retrieve import retrieve_results
from ai_server.retrieve.step4_retrieve import retrieve_results
import json
import asyncio
import re
async def create_json_to_excel(pdf_path, collection_name):
    context_queries, product_keys = await retrieve_results(pdf_path, collection_name)
    print("=== Bắt đầu tạo file Excel ===")
    wb = create_excel_file(context_queries, product_keys)
    print("=== Đã tạo file Excel thành công ===")
    return wb


def create_excel_file(context_queries, product_keys):
    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "BẢNG TUYÊN BỐ ĐÁP ỨNG"
    
    # Thiết lập tiêu đề chính
    ws.merge_cells('A1:F1')
    ws['A1'] = "BẢNG TUYÊN BỐ ĐÁP ỨNG VỀ KỸ THUẬT"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header của bảng
    headers = [
        "Hạng mục số", "Tên hàng hoá",
        "Thông số kỹ thuật và các tiêu chuẩn của hàng hoá trong E-HSMT",
        "Thông số kỹ thuật và các tiêu chuẩn của hàng hoá chào trong E-HSDT",
        "Hồ sơ tham chiếu", "Tình đáp ứng của hàng hoá"
    ]
    
    # Ghi header vào dòng 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Thiết lập borders cho header
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Áp dụng border cho header
    for col in range(1, 7):
        ws.cell(row=2, column=col).border = thin_border
    
    # Ghi dữ liệu
    current_row = 3
    product_counter = 1  # Đếm số La Mã
    
    for product, hang_hoa_dict in product_keys.items():
        # Tạo hàng cho product với số La Mã
        roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", 
                         "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX"]
        
        # Ghi số La Mã vào cột đầu
        ws.cell(row=current_row, column=1).value = roman_numerals[product_counter - 1] if product_counter <= len(roman_numerals) else str(product_counter)
        
        # Merge các ô từ cột 2 đến cột 6 để ghi tên product
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws.cell(row=current_row, column=2).value = product
        
        # Thiết lập style cho hàng product
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        current_row += 1
        product_counter += 1
        
        # Ghi các hàng hàng hóa con
        for hang_hoa_idx, (ten_hang_hoa, items) in enumerate(hang_hoa_dict.items(), start=1):
            ma_ids = items[:-2]  # Các ID
            dap_ung = items[-2]  # ví dụ: "đáp ứng"
            if '/' in dap_ung:
                numerator, denominator = dap_ung.split('/')
                try:
                    numerator = int(numerator)
                    denominator = int(denominator)
                    if denominator == 0 or numerator == 0:  # mẫu bằng 0
                        dap_ung = "Không đáp ứng"
                    elif numerator == denominator:
                        dap_ung = f"Đáp ứng"
                    else:
                        dap_ung = f"Đáp ứng : {numerator} / {denominator}"
                except ValueError:  # tử hoặc mẫu có chữ (không phải số)
                    dap_ung = "Không đáp ứng"
            else:
                try:
                    dap_ung = int(dap_ung)
                    if dap_ung == 0:  # nếu là số thập phân
                        dap_ung = "Không đáp ứng"
                    else:
                        dap_ung = f"Đáp ứng"
                except ValueError:
                    dap_ung = "Không đáp ứng"

            # Lấy danh sách các yêu cầu
            spec_list = []
            dap_ung_list = []
            ho_so_list = []
            page_list = []

            for ma in ma_ids:
                if ma in context_queries:
                    spec_list.append(context_queries[ma]['yeu_cau_ky_thuat_chi_tiet'])
                    # Kiểm tra kha_nang_dap_ung có chứa số hoặc chữ không
                    kha_nang_dap_ung = context_queries[ma].get('kha_nang_dap_ung', '')
                    if kha_nang_dap_ung and re.search(r'[a-zA-Z0-9]', kha_nang_dap_ung):
                        dap_ung_list.append(kha_nang_dap_ung)
                    else:
                        dap_ung_list.append('')
                    ho_so_tham_khao = context_queries[ma].get("tai_lieu_tham_chieu")
                    content = ""
                    file = ho_so_tham_khao.get("file", "")
                    if file:
                        content = f"Tài liệu tham chiếu: {file}"
                    section = ho_so_tham_khao.get("section", "")
                    # if section:
                    #     content += f" - Mục: {section}"
                    table_or_figure = ho_so_tham_khao.get("table_or_figure", "")
                    if table_or_figure is not None and table_or_figure != "":
                        content += f" - Có trong bảng/hình: {table_or_figure}"
                    content += f" - Trang {ho_so_tham_khao.get('page', '')}: \n{ho_so_tham_khao.get('evidence', '')}"

                    ho_so_list.append(content)
                    page_list.append(int(ho_so_tham_khao.get('page', '')))

            # Lưu vị trí bắt đầu để merge sau
            start_row = current_row
            
            # Ghi từng spec thành một hàng riêng biệt
            for spec_idx, (spec, dap_ung_item, ho_so,page) in enumerate(zip(spec_list, dap_ung_list, ho_so_list,page_list)):
                # Ghi dữ liệu vào các cell
                if spec_idx == 0:
                    # Hàng đầu tiên: ghi số thứ tự hàng hóa
                    ws.cell(row=current_row, column=1).value = str(hang_hoa_idx)
                else:
                    # Các hàng tiếp theo: để trống cột 1
                    ws.cell(row=current_row, column=1).value = ""
                
                # Ghi tên hàng hóa tạm thời (sẽ merge sau)
                ws.cell(row=current_row, column=2).value = ten_hang_hoa if spec_idx == 0 else ""
                
                ws.cell(row=current_row, column=3).value = spec
                if dap_ung_item:
                    dap_ung_item = re.sub(r'[\x00-\x1F\x7F]', '', dap_ung_item)  # Loại bỏ ký tự không hợp lệ
                    ws.cell(row=current_row, column=4).value = dap_ung_item

                    ho_so = re.sub(r'[\x00-\x1F\x7F]', '', ho_so)  # Loại bỏ ký tự không hợp lệ
                    # Ghi tài liệu tham chiếu cho từng hàng (không merge)
                    ws.cell(row=current_row, column=5).value = ho_so
                
                # Ghi tình trạng đáp ứng tạm thời (sẽ merge sau)
                ws.cell(row=current_row, column=6).value = dap_ung if spec_idx == 0 else ""
                
                # Thiết lập style cho dòng dữ liệu
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = thin_border
                
                current_row += 1
            
            # Merge các ô cho cột A, B, F (Hạng mục số, Tên hàng hóa, Tình trạng đáp ứng)
            if len(spec_list) > 1:
                end_row = current_row - 1
                
                # Merge cột A (Hạng mục số)
                ws.merge_cells(f'A{start_row}:A{end_row}')
                ws.cell(row=start_row, column=1).value = str(hang_hoa_idx)
                ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Merge cột B (Tên hàng hóa)
                ws.merge_cells(f'B{start_row}:B{end_row}')
                ws.cell(row=start_row, column=2).value = ten_hang_hoa
                ws.cell(row=start_row, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Merge cột F (Tình trạng đáp ứng)
                ws.merge_cells(f'F{start_row}:F{end_row}')
                ws.cell(row=start_row, column=6).value = dap_ung
                ws.cell(row=start_row, column=6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Thiết lập độ rộng cột
    column_widths = {
        'A': 12,  # Hạng mục số
        'B': 25,  # Tên hàng hoá
        'C': 40,  # Thông số E-HSMT
        'D': 40,  # Thông số E-HSDT
        'E': 20,  # Hồ sơ tham chiếu
        'F': 20   # Tình trạng đáp ứng
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Thiết lập chiều cao dòng
    ws.row_dimensions[1].height = 30  # Tiêu đề chính
    ws.row_dimensions[2].height = 60  # Header
    
    # Thiết lập chiều cao tự động cho các dòng dữ liệu
    for row in range(3, current_row):
        ws.row_dimensions[row].height = None  # Auto height
    
    return wb
    # # Lưu file
    # output_path = "D:/study/LammaIndex/output/bang_tuyen_bo_dap_ung10.xlsx"
    # wb.save(output_path)
    # print(f"Đã lưu file Excel tại: {output_path}")
def create_json_to_excel1():
    # context_queries, product_keys = await adapt_or_not(pdf_path, filename_ids, collection_name)
    with open("D:/study/LammaIndex/output/context_queries.json", "r", encoding="utf-8") as f:
        context_queries = json.load(f)
    with open("D:/study/LammaIndex/output/product_keys.json", "r", encoding="utf-8") as f:
        product_keys = json.load(f)
    print("=== Bắt đầu tạo file Excel ===")
    wb = create_excel_file(context_queries, product_keys)
    print("=== Đã tạo file Excel thành công ===")
    output_path = "D:/study/LammaIndex/output/bang_tuyen_bo_dap_ung10.xlsx"
    wb.save(output_path)
    print(f"Đã lưu file Excel tại: {output_path}")

# # Sử dụng
# async def main():
#     await create_json_to_excel("D:/study/LammaIndex/documents/test1.pdf", "hello_my_friend")

# # Fix: Run the async function properly
# if __name__ == "__main__":
#     asyncio.run(main())